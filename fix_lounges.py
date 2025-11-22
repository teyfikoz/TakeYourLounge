#!/usr/bin/env python3
"""
Fix lounges: Remove duplicates, add Excel data, fix country names
"""

import csv
import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Set


# ISO country code to full name mapping
COUNTRY_MAP = {
    'US': 'United States',
    'GB': 'United Kingdom',
    'FR': 'France',
    'DE': 'Germany',
    'ES': 'Spain',
    'IT': 'Italy',
    'TR': 'Turkey',
    'AE': 'United Arab Emirates',
    'SA': 'Saudi Arabia',
    'JP': 'Japan',
    'CN': 'China',
    'IN': 'India',
    'AU': 'Australia',
    'CA': 'Canada',
    'MX': 'Mexico',
    'BR': 'Brazil',
    'SG': 'Singapore',
    'HK': 'Hong Kong',
    'TH': 'Thailand',
    'MY': 'Malaysia',
    'ID': 'Indonesia',
    'PH': 'Philippines',
    'KR': 'South Korea',
    'RU': 'Russia',
    'NL': 'Netherlands',
    'BE': 'Belgium',
    'CH': 'Switzerland',
    'AT': 'Austria',
    'SE': 'Sweden',
    'NO': 'Norway',
    'DK': 'Denmark',
    'FI': 'Finland',
    'PL': 'Poland',
    'CZ': 'Czech Republic',
    'GR': 'Greece',
    'PT': 'Portugal',
    'IE': 'Ireland',
    'NZ': 'New Zealand',
    'ZA': 'South Africa',
    'EG': 'Egypt',
    'IL': 'Israel',
    'QA': 'Qatar',
    'KW': 'Kuwait',
    'OM': 'Oman',
    'BH': 'Bahrain',
    'JO': 'Jordan',
    'LB': 'Lebanon',
    'KE': 'Kenya',
    'NG': 'Nigeria',
    'GH': 'Ghana',
    'AR': 'Argentina',
    'CL': 'Chile',
    'CO': 'Colombia',
    'PE': 'Peru',
    'VE': 'Venezuela',
    'MA': 'Morocco',
    'TN': 'Tunisia',
    'LY': 'Libya',
    'DZ': 'Algeria',
    'ET': 'Ethiopia',
    'UG': 'Uganda',
    'TZ': 'Tanzania',
    'KZ': 'Kazakhstan',
    'UZ': 'Uzbekistan',
    'GE': 'Georgia',
    'AM': 'Armenia',
    'AZ': 'Azerbaijan',
    'BY': 'Belarus',
    'UA': 'Ukraine',
    'RO': 'Romania',
    'BG': 'Bulgaria',
    'RS': 'Serbia',
    'HR': 'Croatia',
    'SI': 'Slovenia',
    'SK': 'Slovakia',
    'HU': 'Hungary',
    'LT': 'Lithuania',
    'LV': 'Latvia',
    'EE': 'Estonia',
    'IS': 'Iceland',
    'LU': 'Luxembourg',
    'MT': 'Malta',
    'CY': 'Cyprus',
    'VN': 'Vietnam',
    'LA': 'Laos',
    'KH': 'Cambodia',
    'MM': 'Myanmar',
    'BD': 'Bangladesh',
    'LK': 'Sri Lanka',
    'NP': 'Nepal',
    'PK': 'Pakistan',
    'AF': 'Afghanistan',
    'IR': 'Iran',
    'IQ': 'Iraq',
    'SY': 'Syria',
    'YE': 'Yemen',
    'MV': 'Maldives',
    'MU': 'Mauritius',
    'SC': 'Seychelles',
    'RE': 'Reunion',
    'MG': 'Madagascar',
    'ZW': 'Zimbabwe',
    'ZM': 'Zambia',
    'MW': 'Malawi',
    'MZ': 'Mozambique',
    'NA': 'Namibia',
    'BW': 'Botswana',
    'LS': 'Lesotho',
    'SZ': 'Eswatini',
    'AO': 'Angola',
    'CG': 'Congo',
    'CD': 'Democratic Republic of Congo',
    'GA': 'Gabon',
    'CM': 'Cameroon',
    'CI': 'Ivory Coast',
    'SN': 'Senegal',
    'ML': 'Mali',
    'NE': 'Niger',
    'BF': 'Burkina Faso',
    'TG': 'Togo',
    'BJ': 'Benin',
    'GW': 'Guinea-Bissau',
    'GN': 'Guinea',
    'SL': 'Sierra Leone',
    'LR': 'Liberia',
    'GM': 'Gambia',
    'MR': 'Mauritania',
}


class LoungeFixer:
    """Fix and merge lounge data"""

    def __init__(self):
        self.lounges: Dict[str, Dict] = {}
        self.duplicates_removed = 0

    def normalize_name(self, name: str, airport: str = '') -> str:
        """Create normalized key for duplicate detection"""
        # Remove common prefixes/suffixes
        name = name.lower().strip()
        name = name.replace('lounge', '').replace('vip', '').replace('the', '')
        name = name.strip()
        airport = airport.lower().strip()
        return f"{airport}_{name}"

    def fix_country_name(self, country: str, iso_country: str = '') -> str:
        """Convert ISO code to full country name"""
        if not country or len(country) == 2:
            # It's an ISO code
            return COUNTRY_MAP.get(country or iso_country, country or iso_country)
        # Already full name
        return country

    def load_existing_lounges(self, csv_path: Path):
        """Load existing lounges from CSV"""
        print(f"\n📊 Loading existing lounges from {csv_path.name}...")

        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Create unique key
                key = self.normalize_name(row.get('name', ''), row.get('airport_code', ''))

                # Fix country name
                row['country'] = self.fix_country_name(
                    row.get('country', ''),
                    row.get('iso_country', '')
                )

                if key in self.lounges:
                    self.duplicates_removed += 1
                else:
                    self.lounges[key] = row

        print(f"   ✅ Loaded {len(self.lounges)} unique lounges")
        print(f"   🗑️  Removed {self.duplicates_removed} duplicates")

    def add_excel_lounges(self, excel_path: Path):
        """Add lounges from Excel file"""
        print(f"\n📊 Loading Excel lounges from {excel_path.name}...")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        added = 0
        skipped = 0

        for row_num in range(2, ws.max_row + 1):
            name = ws.cell(row_num, 4).value  # Lounge Name
            airport = ws.cell(row_num, 2).value  # Airport Name
            city = ws.cell(row_num, 3).value
            country = ws.cell(row_num, 6).value
            terminal = ws.cell(row_num, 7).value
            region = ws.cell(row_num, 5).value

            if not name or not airport:
                continue

            # Create key
            key = self.normalize_name(name, airport)

            if key not in self.lounges:
                # Add new lounge
                lounge_id = f"global_{row_num}_{key}".replace(' ', '_').replace(',', '')[:100]

                self.lounges[key] = {
                    'id': lounge_id,
                    'name': name,
                    'airport_code': '',
                    'airport_name': airport,
                    'city': city or '',
                    'country': country or '',
                    'terminal': terminal or '',
                    'location': '',
                    'open_hours': '',
                    'amenities': '["Free Wi-Fi", "Food & Beverages", "Comfortable Seating"]',
                    'access_methods': '["Day Pass"]',
                    'lounge_type': 'independent',
                    'description': f'{name} at {airport}, {city}.',
                    'conditions': '',
                    'rating': 0.0,
                    'review_count': 0,
                    'max_capacity': 50,
                    'coordinates': '',
                    'latitude': '',
                    'longitude': '',
                    'elevation_ft': '',
                    'continent': region or '',
                    'iso_country': '',
                    'wikipedia_link': '',
                    'images': '[]',
                    'source': 'GlobalLoungeList',
                    'created_at': ''
                }
                added += 1
            else:
                skipped += 1

        print(f"   ✅ Added {added} new lounges from Excel")
        print(f"   ⏭️  Skipped {skipped} duplicates")

    def export_fixed_csv(self, output_path: Path):
        """Export fixed CSV"""
        print(f"\n💾 Exporting fixed CSV to {output_path}...")

        fieldnames = [
            'id', 'name', 'airport_code', 'airport_name', 'city', 'country',
            'terminal', 'location', 'open_hours', 'amenities', 'access_methods',
            'lounge_type', 'description', 'conditions', 'rating', 'review_count',
            'max_capacity', 'coordinates', 'latitude', 'longitude', 'elevation_ft',
            'continent', 'iso_country', 'wikipedia_link', 'images', 'source', 'created_at'
        ]

        with open(output_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()

            for lounge in sorted(self.lounges.values(), key=lambda x: (x.get('country', ''), x.get('city', ''), x.get('name', ''))):
                writer.writerow(lounge)

        print(f"   ✅ Exported {len(self.lounges)} lounges")

    def print_stats(self):
        """Print statistics"""
        countries = set(l.get('country') for l in self.lounges.values() if l.get('country'))
        cities = set(l.get('city') for l in self.lounges.values() if l.get('city'))

        print("\n" + "="*60)
        print("📊 FINAL STATISTICS")
        print("="*60)
        print(f"\n✅ Total Unique Lounges: {len(self.lounges)}")
        print(f"🗑️  Duplicates Removed: {self.duplicates_removed}")
        print(f"🌍 Countries: {len(countries)}")
        print(f"🏙️  Cities: {len(cities)}")
        print("\n" + "="*60)


def main():
    print("""
╔═══════════════════════════════════════════════════════════╗
║              LOUNGE DATA FIXER                            ║
║     Remove Duplicates + Fix Countries + Add Excel        ║
╚═══════════════════════════════════════════════════════════╝
    """)

    data_dir = Path(__file__).parent
    input_csv = data_dir / "master_lounges_enriched.csv"
    excel_path = Path("/Users/teyfikoz/Desktop/Tech Sync Analytica/Global Lounge List.xlsx")
    output_csv = data_dir / "master_lounges_fixed.csv"

    if not input_csv.exists():
        print(f"❌ Error: {input_csv} not found!")
        return

    if not excel_path.exists():
        print(f"⚠️  Warning: {excel_path} not found! Continuing without Excel data...")
        excel_path = None

    # Fix lounges
    fixer = LoungeFixer()
    fixer.load_existing_lounges(input_csv)

    if excel_path:
        fixer.add_excel_lounges(excel_path)

    fixer.export_fixed_csv(output_csv)
    fixer.print_stats()

    print(f"\n✅ Lounge fixing complete!")
    print(f"📁 Output: {output_csv}")
    print("\n💡 Next steps:")
    print("   1. Review the fixed CSV")
    print("   2. Regenerate web data: python3 generate_web_data.py")
    print("   3. Refresh browser to see changes")


if __name__ == "__main__":
    main()
